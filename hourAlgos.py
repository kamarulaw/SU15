import os
import datetime 
import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


#GLOBAL VARIABLES 
NROWS = 1054 #number of rows in the .xlsx file. function to set value will be made later"
FNAME = "HoursbyPerson_Ade_v2.xlsx"

HOURCOL = 6
LNAMECOL = 1
FNAMECOL = 2
PIDCOL = 4
LEVELVAL = 3
DATEROW = 5

#GLOBAL SETTERS
def setter():
    global FNAME,LNAMECOL,FNAMECOL,PIDCOL,HOURCOL,NROWS,LEVELVAL,DATEROW 
    FNAME = str(raw_input("File: "))
    LNAMECOL = int(raw_input("Last Name Column: "))
    FNAMECOL = int(raw_input("First Name Column: "))
    PIDCOL = int(raw_input("Project ID Column: "))
    DATEROW = int(raw_input("Hours Date Column: "))
    HOURCOL = int(raw_input("Entered Hours Column: "))
    NROWS = int(raw_input("Number of Rows in Excel File:"))
    LEVELVAL = int(raw_input("Project Level Depth:"))+1

def hoursByProjectLevel(filename=FNAME,level=LEVELVAL,rows=NROWS):
    #Workbook being read from
    tsbook = load_workbook(filename)
    tsheet = tsbook.active
    namearr = []
    counter=0
    pfiles = [] #keeps track of the number of project ids
    mdates = [] #keeps track of the number of mm/yy 
    for row in range(2,rows+1):#rows+1):
        name = createName(row,tsheet) #name for row
        plevel = createProject(row,tsheet,level) #project for row
        nindex = nameFound(name,namearr) #location of name in array
        darray = tsheet.cell(row=row,column=DATEROW).value.date().strftime("%m/%d/%y").split("/",3)
        mDate = createMDate(darray)
        if mDate not in mdates:
            mdates.append(mDate)
        if plevel not in pfiles:
            pfiles.append(plevel)
        if (nindex == -1):
            namearr.append([name,[]])
            nsize = len(namearr)
            namearr[nsize-1][1].append([plevel,[[createMDate(darray),tsheet.cell(row=row,column=HOURCOL).value]]])   
        else:
            loc = projectFound(plevel,namearr[nindex][1])
            datePresent = mDateFound(darray,namearr[nindex][1][0][1])
            if (loc==-1):
                namearr[nindex][1].append([plevel,[[createMDate(darray),tsheet.cell(row=row,column=HOURCOL).value]]])
            else:
                if (datePresent==-1):
                    namearr[nindex][1][0][1].append([createMDate(darray),tsheet.cell(row=row,column=HOURCOL).value])
                else:
                    namearr[nindex][1][0][1][datePresent][1]+=tsheet.cell(row=row,column=HOURCOL).value 
    outputData(nameSort(namearr),dateSort(mdates),pfiles)
    return nameSort(namearr)

def outputData(arr,mdates,pfiles):
    #Workbook being written to
    wb = Workbook()
    wsheet = wb.active
    wsheet.title = "Results"
    wsheet["A3"] = "Summary"
    numberoftimes = len(mdates)
    numberofprojects = len(pfiles)
    rtracker = [3,True] # rtracker: info about the row [row,isFilenameRow]
    ctracker = [3,0] #ctracker: infor about the column [column]

    #horizontal output
    for elem in mdates:
        wsheet.cell(row=1,column=ctracker[0]).value = elem
        wsheet.cell(row=2,column=ctracker[0]).value = "Plan"
        ctracker[0]+=1

        wsheet.cell(row=1,column=ctracker[0]).value = elem
        wsheet.cell(row=2,column=ctracker[0]).value = "Actual"
        ctracker[0]+=1

        wsheet.cell(row=1,column=ctracker[0]).value = elem
        wsheet.cell(row=2,column=ctracker[0]).value = "Variance"
        ctracker[0]+=2        

    ctracker[0] = 4 #the first "Time/Actual" column
    #vertical output 
    for i in range(len(arr)):
        cell = "B" + str(rtracker[0])
        wsheet[cell] =  arr[i][0]
        for i in range(len(mdates)):
            wsheet.cell(row=rtracker[0],column=ctracker[0]).value = personSum(arr,wsheet[cell].value,mdates[i],"")
            #print wsheet[cell].value
            ctracker[0]+=4
        ctracker[0]=4 #goes back to first "Time/Actual" col row iteration
        rtracker[0]+=1
        
    rtracker[0]+=1
    
    for elem in pfiles:
        for i in range(len(arr)):
            cell = "B"+str(rtracker[0])
            if (rtracker[1]==True):
                wsheet["A"+str(rtracker[0])] = elem
                rtracker[1]=False
            wsheet["B"+str(rtracker[0])] = arr[i][0]
            for i in range(len(mdates)):
                wsheet.cell(row=rtracker[0],column=ctracker[0]).value = personSum(arr,wsheet[cell].value,mdates[i],elem)
                ctracker[0]+=4
            rtracker[0]+=1
            ctracker[0]=4
        rtracker[1]=True
        rtracker[0]+=1
    wb.save("resultfile.xlsx")


    
def mDateFound(darray,arr):
    for i in range(len(arr)):
        elem = arr[i][0]
        tmp = elem.split("/",2)
        if (darray[0]==tmp[0] and darray[2]==tmp[1]):
            return i
    return -1

def createMDate(dayarr):
    return dayarr[0]+"/"+dayarr[2]

def createName(r,tsheet):
    #creates the name that will be used to represent the people in the first two columns
    rname = str(tsheet.cell(row=r,column=LNAMECOL).value)+","+str(tsheet.cell(row=r,column=FNAMECOL).value).replace(" ","")
    return rname.lower()

def createProject(r,tsheet,level):
    #creates the project number as a string for an arbitrary row
    u = tsheet.cell(row=r,column=PIDCOL).value
    pid = str(u)
    levelarr = pid.split(".",5)
    plevel=""
    for i in range(level):
        plevel+=levelarr[i]
        if (i==level-1):
            break 
        plevel+="."
    return plevel

def nameFound(name,arr):
    for i in range(len(arr)):
        if name==arr[i][0]:return i
    return -1
    
def projectFound(p,arr):
    for elem in arr:
        if p==elem[0]:return arr.index(elem)
    return -1

def displayInfo(arr):
    for elem in arr:
        print elem
        print 

def nameSort(arr):
    for i in range(len(arr)):
        minimum = i
        for k in range(i+1,len(arr)):
            if arr[k][0] < arr[minimum][0]: 
                minimum=k
        swap(arr,minimum,i)
    return arr

def dateSort(arr):
    for i in range(len(arr)):
        minimum = i
        for k in range(i+1,len(arr)): #lexicographic function will go here
            if dateLessThan(arr[k],arr[minimum]):
                minimum = k
        swap(arr,minimum,i)
    return arr

def dateLessThan(d1,d2):
    arr1 = d1.split("/",2)
    arr2 = d2.split("/",2)
    if int(arr1[1]) < int(arr2[1]):
        return True
    elif int(arr1[1] > arr2[1]):
        return False
    else:
        if arr1[0]<arr2[0]:
            return True
        else:
            return False
        
def swap(arr,i,j):
    temp = arr[j]
    arr[j] = arr[i]
    arr[i] = temp
    return

def personSum(arr,person,date,pmode):
    nindex = nameFound(person,arr)
    total = 0
    if nindex==-1:
        return 0
    else:
        if pmode=="":
            for hourset in arr[nindex][1]:
                for vals in hourset[1]:
                    if (vals[0]==date):
                            total+=vals[1]
        else:
            for proj in arr[nindex][1]:
                if proj[0]==pmode:
                    for vals in proj[1]:
                        if date==vals[0]:
                            total+=vals[1]
    return total


######### TEST CODE #########

def hoursByName(filename=FNAME,name="woodford,paul",rows=NROWS):
    #THIS FUNCTION IS MEANT TO TEST THAT HOURSBYPROJECTLEVEL IS WORKING 
    """
    inputs
    - xlsx filename: if it is a csv file then see the link below
    http://bit.ly/1LRFo58 (converting csv to xlsx)

    - name: name of the employee should be lastname ane then firstname with no spaces
    between

    - rows: the number of rows in the table
    """
    tsbook = load_workbook(filename)
    tsheet = tsbook.active
    hours = 0
    for i in range(2,rows+1):
        rname = str(tsheet.cell(row=i,column=1).value)+","+str(tsheet.cell(row=i,column=2).value)
        if (name.lower() == rname.lower()):
            hours+=tsheet.cell(row=i,column=6).value
    return hours 

#Run Code
setter()

displayInfo(hoursByProjectLevel(filename=FNAME,rows=NROWS,level=LEVELVAL))
#print hoursByProjectLevel(filename=FNAME,rows=NROWS,level=LEVELVAL)
#print hoursByName(filename=FNAME,name="marshall,aaron",rows=NROWS)
#print personSum(hoursByProjectLevel(filename=FNAME,rows=NROWS,level=LEVELVAL),'davieau,gerald',"08/15","023116")
arr = hoursByProjectLevel(filename=FNAME,rows=NROWS,level=LEVELVAL)

